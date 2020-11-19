#############
# LIBRARIES #
#############

# General
import os
import re
import requests
import sys
import yaml
from typing import Any, Dict, List, TypedDict

# Sheet creation
import csv
import datetime
import openpyxl

# Email
import email.mime.application
import email.mime.base
import email.mime.multipart
import email.mime.text
import getpass
import smtplib
import ssl

##########
# CONFIG #
##########

# Sheet
csv_name = os.path.abspath(sys.argv[1]) # Get filename via CMD input
headers_german = ["Datum", "Anfang", "Ende", "anwesend", "Pausen", "gearbeitet", "Kommentare"]

# Personal
weekly_worktime = "15:00"

# Email
sender_email = "julius.schlensok@mitsm.de"
receiver_email = "office@mitsm.de"
email_body = "Dies ist eine automatisierte Email, an die Julius' Arbeitszeit angehängt wird. Probleme/Feedback bitte an feedback@schlensok.org"
smtp_server = "smtp.office365.com"
name = "Julius Schlensok"

###############
# DEFINITIONS #
###############



def parseCsv(filename: str) -> List[List[str]]:
	results = []
	# Read file
	with open(filename) as csvfile:
		reader = csv.reader(csvfile, delimiter=',')
		for row in reader:
			results.append(row)
	# Rearrange columns
	del results[0] # Remove header row (we're using our own headerns)
	results = list(map(list, zip(*results))) # transpose
	results[4], results[5], results[6] = results[6], results[4], results[5] # move "breaks" left of "rel. duration" and "description"
	del results[-1] # remove "break description" column (automatically exported with break duration)
	
	return results


class Timesheet:

	def __init__(self):
		self.wb = openpyxl.Workbook()
		self.sheet = self.wb.active
		self.styles = {}
		self.loadStyles(self.styles)

	def loadStyles(self, styles: Dict[str, openpyxl.styles.NamedStyle]) -> None:
		styles["duration"] = openpyxl.styles.NamedStyle(name="duration_style", number_format="HH:MM:SS")
		styles["time"] = openpyxl.styles.NamedStyle(name="time_style", number_format="HH:MM")
		styles["header"] = openpyxl.styles.NamedStyle(name="header_style")
		styles["header"].font = openpyxl.styles.Font(bold=True)
		styles["date"] = openpyxl.styles.NamedStyle(name="date_style", number_format="DD.MM.YYYY")
		
		# Multiple border styles
		styles["border_lines"] = {"thin": openpyxl.styles.borders.Side(border_style="thin")}

	def setHeaderRow(self, headers: List[str]) -> None:
		for i in range(len(headers)):
			cell = self.sheet.cell(row=1, column=i+1)
			cell.value = headers[i]
			cell.style = self.styles["header"]

	def setDates(self, dates: List[str]) -> None:
		dates = [datetime.datetime.strptime(x, "%d.%m.%Y") for x in dates] # convert times from string to datetime objects
		for i in range(len(dates)):
			cell = self.sheet.cell(column=1, row=i+2) # first column, starting from second row
			cell.value = dates[i]
			cell.style = self.styles["date"]

	def setTimes(self, times: [List[str], List[str]]) -> None:
			times = [[datetime.datetime.strptime(x, "%d.%m.%Y %I:%M %p") for x in time] for time in times] # convert times from strings to datetime objects
			for i in range(len(times)):
				for j in range(len(times[i])):
					cell = self.sheet.cell(row=j+2, column=i+2) # second/third column, starting from second row
					cell.value = times[i][j]
					cell.style = self.styles["time"]

	def setBreaks(self, times: List[str]) -> None:
		times = [datetime.timedelta(hours=x.hour, minutes=x.minute) for x in [datetime.datetime.strptime(y, "%H:%M:%S") for y in times]]
		for i in range(len(times)):
			cell = self.sheet.cell(column=5, row=i+2) # 5th column, starting with second entry
			cell.value = times[i]
			cell.style = self.styles["duration"]

	def calculateWorktime(self) -> None:
		for i in range(len(self.sheet['A'])-1):
			if self.sheet.cell(column=2, row=i+2).value: # if not holiday
				cell = self.sheet.cell(column=4, row=i+2) # 4th column, starting with second entry
				cell.value = f"=C{i+2}-B{i+2}"
				cell.style = self.styles["duration"]

	def subtractBreaks(self) -> None:
		for i in range(len(self.sheet['A'])-1):
			if self.sheet.cell(column=2, row=i+2).value: # if not holiday
				cell = self.sheet.cell(column=6, row=i+2) # 7th column, starting with second entry
				cell.value = f"=D{i+2}-E{i+2}"
				cell.style = self.styles["duration"]

	def collapseDays(self) -> None:
		# Iterate over all row-higher row combinations. The list comprehensions are required as iter_rows() otherwise returns single cells in tuples.
		for cell1 in [x[0] for x in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, min_col=1, max_col=1)]:
			first_row = cell1.row
			collapse = False
			for cell2 in [x[0] for x in self.sheet.iter_rows(min_row=cell1.row+1, max_row=self.sheet.max_row, min_col=1, max_col=1)]:
				if cell1.value == cell2.value: # same date
					last_row = cell2.row
					if not collapse:
						collapse = True
						# Initialize totals with first session's values
						breaktime = self.sheet[f"E{first_row}"].value
						day_start = self.sheet[f"B{first_row}"].value
					day_end = self.sheet[f"C{last_row}"].value # set day's end to last row's
					print(self.sheet[f"B{last_row}"].value)
					print(self.sheet[f"C{last_row-1}"].value)
					breaktime += (self.sheet[f"B{last_row}"].value - self.sheet[f"C{last_row-1}"].value) # increase breaktime by difference between two sessions
					breaktime += self.sheet[f"E{last_row}"].value # increase breaktime by session's
				elif collapse: # all entries for the day found, let's collapse it
					self.sheet[f"C{first_row}"].value = day_end # set end time
					self.sheet[f"E{first_row}"].value = breaktime
					self.sheet.delete_rows(first_row+1, last_row-first_row) # Delete all sessions after the first
					collapse = False
				else:
					continue

	def sumAndFormatWorktime(self, background_color="00FFFF"):
		max_row = self.sheet.max_row
		# Insert sum function
		sum_cell = self.sheet.cell(row=max_row+1, column= 6) # Cell below effective worktime column
		sum_cell.value = f"=SUM(F2:F{max_row})" # insert a simple Excel function
		sum_cell.number_format = openpyxl.styles.numbers.FORMAT_DATE_TIMEDELTA # so it doesn't overflow after 24 hours
		# Insert "SUM" text cell
		sum_desc_cell = self.sheet.cell(row=max_row+1, column=5)
		sum_desc_cell.value = "SUMME"
		sum_desc_cell.style = self.styles["header"]
		# Set background
		for row in range(1, max_row+2):
			self.sheet.cell(row=row, column=6).fill = openpyxl.styles.PatternFill(fgColor=background_color, fill_type="solid")
		# Create box around cells
		self.drawBorderLines(self.generateCellRangeString(sum_cell, sum_desc_cell)) # Draw border around these two cells
		
	def generateBorders(self, style: str, *args) -> openpyxl.styles.Border:
		"""Helper function that returns a Border object with lines on the sides specified
		"""
		borders = {"top": None, "right": None, "bottom": None, "left": None}
		for arg in args:
			borders[arg] = self.styles["border_lines"][style]

		return openpyxl.styles.Border(**borders)

	def drawBorderLines(self, cell_range: str) -> None:
		"""Helper function so we can differentiate between cases for the main one
		"""
		assert re.compile(r"^[A-Z]+[0-9]+:[A-Z]+[0-9]+$").match(cell_range) # range should look like A4:C7.
		
		# Credit to Rabih Kodeih on StackOverflow
		rows = self.sheet[cell_range]
		if len(rows) == 1 and len(rows[0]) == 1: # just one cell
			rows[0][0].border = self.generateBorders("thin", "top", "right", "bottom", "left")

		elif len(rows) == 1 and len(rows[0]) > 1: # horizontal vector (just one row)
			rows[0][0].border = self.generateBorders("thin", "top", "left", "bottom") # left end cell
			rows[0][-1].border = self.generateBorders("thin", "top", "right", "bottom") # right end cell
			if len(rows[0]) > 2: # cells in between
				for cell in rows[0][1:-1]:
					cell.border = self.generateBorders("thin", "top", "bottom")

		elif len(rows) > 1 and len(rows[0]) == 1: # vertical vector (just one column)
			rows[0][0].border = self.generateBorders("thin", "left", "top", "right") # top end cell
			row[-1][0].border = self.generateBorder("left", "bottom", "right") # bottom end cell
			if len(rows) > 2: # cells in between
				for row in rows[1:-1]:
					row[0].border = self.generateBorders("thin", "left", "right")

		else: # multiple rows & columns
			# Corners
			rows[0][0].border = self.generateBorders("thin", "top", "left")
			rows[0][-1].border = self.generateBorders("thin", "top", "right")
			rows[-1][-1].border = self.generateBorders("thin", "bottom", "right")
			rows[-1][0].border = self.generateBorders("thin", "bottom", "left")

			for row in rows[1:-1]: # iterate down through rows and set leftmost and rightmost cell's borders
				row[0].border = self.generateBorders("thin", "left")
				row[-1].border = self.generateBorders("thin", "right")
			for col in rows[0][1:-1]: # set top border (first row)
				col.border = self.generateBorders("thin", "top")
			for col in rows[-1][1:-1]: # set bottom border (last row)
				col.border = self.generateBorders("thin", "bottom")
	
	def generateCellRangeString(self, cell1: openpyxl.cell.cell.Cell, cell2: openpyxl.cell.cell.Cell) -> str:
		"""Takes two cells and draws a border around the the area in between them
		"""
		if cell1.column <= cell2.column and cell1.row <= cell2.row: # Top-left and bottom-right cell
			cell_range = ':'.join([f"{chr(ord('A') + cell.column - 1)}{cell.row}" for cell in [first_cell, last_cell]])
		else: # Find out min/max cols/rows
			borders = {}
			borders["left"], borders["right"] = chr(ord('A') + min(cell1.column, cell2.column) -1), chr(ord('A') + max(cell1.column, cell2.column) -1) # Convert column numbers into letters
			borders["top"], borders["bottom"] = min(cell1.row, cell2.row), max(cell1.row, cell2.row)
			cell_range = f"{borders['left']}{borders['top']}:{borders['right']}{borders['bottom']}"
		return cell_range

	def save(self, filename: str=None, overwrite=False) -> None:
		"""
		"""
		filename = self.output_name if not filename else filename
		if overwrite:
			try: # file exists
				os.remove(filename)
			except OSError: # file doesn't exist
				pass
		self.wb.save(filename=filename)

	def confirm(self) -> bool:
		"""
		"""
		os.startfile(f'\"{self.output_name}\"') # Open generated file in Excel. Double quotes around the name because it contains whitespace & special characters (/)
		print("Opening up generated spreadsheet in Excel, this might take a couple of seconds...\n")
		while True:
			confirmation = input("If you want to send the generated file, type: y\nIf you want to quit, type: n\nIf you want to modify the file in Excel: edit and save it (under the same name), then press y\nWaiting...\n").lower()
			if confirmation in ["y", "yes"]:
				return True
				break
			elif confirmation in ["n", "no"]:
				return False
				break
			else:
				print("Please specify yes/no")
				continue # If no valid answer is entered, ask again

	def insertDates(self, dates: List[TypedDict("AdditionalDate", {"name": str, "date": datetime.datetime, "duration": datetime.timedelta})]) -> None:
		"""
		assumption: holidays are ordered by date in increasing order
		"""
		lowest_date_index = 2
		for day in dates:
			inserted = False # to make sure holiday dates are inserted even if no date is larger than them
			for row_num in range(lowest_date_index, self.sheet.max_row+1): # Iterate over all rows and compare dates
				if self.sheet.cell(row=row_num, column=1).value.day > day["date"].day:
					# insertion position found: next row's date is higher than holiday's
					insertion_index = row_num
					inserted = True
					break

			if not inserted: # holiday date is larger than all workday dates
					insertion_index = len(self.sheet['A'])+1 # append

			# Set date, duration, and comment
			self.sheet.insert_rows(insertion_index)
			self.sheet.cell(row=insertion_index, column=1).value = day["date"]
			self.sheet.cell(row=insertion_index, column=1).style = self.styles["date"]
			self.sheet.cell(row=insertion_index, column=6).value = day["duration"]
			self.sheet.cell(row=insertion_index, column=6).style = self.styles["duration"]
			self.sheet.cell(row=insertion_index, column=7).value = day["name"]

	def setOutputName(self, name: str, filetype: str="xlsx") -> None:
		"""
		"""
		first_date = self.sheet["A2"].value
		last_date = self.sheet[f"A{self.sheet.max_row-1}"].value
		print(first_date)
		print(last_date)
		print(first_date == last_date)
		span = False

		# Set year(s)
		if first_date.year == last_date.year:
			year = first_date.year
		else: # timesheet spans multiple years
			span = True

		# Set month(s)
		# Case A: Timesheet spans just one month
		if first_date.month == last_date.month:
			month = first_date.month
		# Case B: Timesheet spans multiple months
		else:
			first_month = first_date.month
			last_month = last_date.month
			# Case B1: Timesheet spans a quarter
			if first_month in [1, 4, 7, 10] and last_month-first_month==2: # month span is quarter
				month = f"Q{int(first_month/3)+1}" # set quarter name
			# Case B2: Timesheet spans multiple months across different years
			elif first_year := first_date.year != (last_year := last_date.year):
				month = f"{first_month} {first_year} - {last_month} {last_year}"
			# Case B3: Timesheet spans multiple months of same year
			else:
				month = f"{first_month}-{last_month}"

		self.output_name = f"Arbeitszeit {name} {month} {year}.{filetype}"

	def fitWidth(self, column: str) -> None:
		length = max(len(str(cell.value)) for cell in self.sheet[column])
		self.sheet.column_dimensions[column].width = length


class Email:
	def __init__(self, from_email, to_email):
		self.message = email.mime.multipart.MIMEMultipart()
		self.message["From"] = self.from_email = from_email
		self.message["To"] = self.to_email = to_email

	def configureEmailServer(self, smtp_server: str, port: int=587) -> None:
		self.port = port
		assert re.compile("([a-z0-9]+\.)+(com|de|org)").match(smtp_server) # Check whether server address is valid
		self.smtp_server = smtp_server

	def setSubject(self, name: str) -> None:
		self.message["Subject"] = subject

	def setSubject(self, name: str, month: int, year: int) -> None:
		month_names = {1: 'Januar', 2: 'Februar', 3: 'März', 4: 'April',	5: 'Mai', 6: 'Juni',
		7: 'Juli', 8: 'August', 9: 'September', 10: 'Oktober', 11: 'November', 12: 'Dezember'
		}
		subject = f"Arbeitszeit {name} {month_names[month]} {year}"
		self.message["Subject"] = subject

	def setSubject(self, subject: str) -> None:
		self.message["Subject"] = subject

	def setBody(self, body: str) -> None:
		self.message.attach(email.mime.text.MIMEText(body, "plain"))

	def passwordPrompt(self) -> None:
		self.password = getpass.getpass("Please enter your email account's password: ")

	def attachFile(self, filename: str) -> None:
		with open(filename, "rb") as file:
			part = email.mime.application.MIMEApplication(file.read(), Name="MIMEApplication")
		part["Content-Disposition"] = f"attachment; filename={filename}"
		self.message.attach(part)

	def send(self) -> None:
		context = ssl.create_default_context()
		text = self.message.as_string()
		with smtplib.SMTP(self.smtp_server, self.port) as server:
			server.ehlo()
			server.starttls(context=context)
			server.login(self.from_email, self.password)
			server.sendmail(self.from_email, self.to_email, text)
			server.quit()


class Holidays:

	def __init__(self):
		self.days = {}

	def fetch(self, month: int = None) -> None:
		"""
		Dict[datetime.datetime, str]
		"""
		today = datetime.date.today() # Get current date for comparison later
		month = today.month if not month else month # if no specific month was requested: use current one

		r = requests.get(f'http://feiertage-api.de/api/?jahr={today.year}&nur_land=BY') # Get holidays for whole year from API
		j = r.json() # Convert response into JSON
		
		# remove holidays not relevant to Munich
		for key in ["Augsburger Friedensfest", "Buß- und Bettag"]:
			del j[key]

		# Convert dates from str to datetime to make it easier to work with them. No idea how to do this with a dict comprehension.
		for key in j.keys():
			j[key]["datum"] = datetime.datetime.strptime(j[key]["datum"], "%Y-%m-%d")

		self.days = {j[key]["datum"]: key for key in j.keys() if j[key]["datum"].month == month}

	def convertToWorktime(self,weekly_worktime: str) -> None:
		dates, names = self.days.keys(), self.days.values()
		assert len(dates) == len(names)
		weekly_worktime = datetime.datetime.strptime(weekly_worktime, "%H:%M") # Convert string to datetime object
		daily_worktime = datetime.timedelta(hours=weekly_worktime.hour, minutes=weekly_worktime.minute) / 5 # Convert into timedelta, and divide by days
		
		self.days = [dict(zip(("name", "date", "duration"), val)) for val in zip(names, dates, [daily_worktime] * len(names))]

	def get(self) -> List[TypedDict("AdditionalDate", {"name": str, "date": datetime.datetime, "duration": datetime.timedelta})]:
		return self.days

###############
# MAIN METHOD #
###############

def main() -> None:
	data = parseCsv(csv_name)

	timesheet = Timesheet()
	timesheet.setHeaderRow(headers_german)
	timesheet.setDates(data[0])
	timesheet.setTimes(data[1:3])
	timesheet.setBreaks(data[4])

	holidays = Holidays()
	holidays.fetch(month=timesheet.sheet["A2"].value.month)
	holidays.convertToWorktime(weekly_worktime)
	timesheet.insertDates(holidays.get())

	# timesheet.collapseDays()
	timesheet.calculateWorktime()
	timesheet.subtractBreaks()
	timesheet.sumAndFormatWorktime()
	timesheet.fitWidth("A")

	timesheet.setOutputName(name)
	timesheet.save(overwrite=True)
	
	confirmed = timesheet.confirm()
	if not confirmed:
		print("Aborting...")
		sys.exit()

	email = Email(sender_email, receiver_email)
	email.configureEmailServer(smtp_server)
	email.setSubject(timesheet.output_name) # Subject = sheet name without file extension
	email.setBody(email_body)
	email.attachFile(timesheet.output_name)
	email.passwordPrompt()
	email.send()

if __name__ == "__main__":
	main()
